import re

from bs4 import BeautifulSoup
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore", category=UserWarning, message="The input looks more like a URL than markup.")

workbook = load_workbook('分类算法数据-师兄.xlsx')
# 选择要读取的工作表
worksheet = workbook['Sheet1']

# 打开要写入的txt文件
with open('output.txt', 'w', encoding='utf-8') as txt_file:
    # 遍历行，读取B和F列的值并写入txt文件
    for row in worksheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始读取数据
        b_column_rich_text = row[1]  # B列的富文本

        # 剔除换行符
        plain_text = re.sub(r'\n', ' ', b_column_rich_text)

        # 可选的：进行文本清洗，去除多余的空格、特殊字符等
        plain_text = re.sub(r'https://\S+', '', plain_text)

        column_f_value = row[5]  # F列的值

        txt_file.write(f'{plain_text}_!_{column_f_value}\n')

# 关闭Excel文件
workbook.close()
